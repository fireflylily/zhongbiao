#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将JSON测试用例导出为Excel文件

用途：
- 将 business_response_test_cases.json 转换为 Excel 格式
- 方便非技术人员查看和编辑测试用例
- 支持批量查看所有测试数据

使用方法：
    python tests/scripts/export_to_excel.py

输出：
    tests/data/business_response_test_cases.xlsx

作者：AI Tender System
日期：2025-12-02
"""

import json
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ 缺少依赖包，请先安装：")
    print("   pip install pandas openpyxl")
    sys.exit(1)


def export_to_excel():
    """将JSON测试用例导出为Excel"""

    # 文件路径
    script_dir = Path(__file__).parent
    json_file = script_dir.parent / "data" / "business_response_test_cases.json"
    excel_file = script_dir.parent / "data" / "business_response_test_cases.xlsx"

    # 检查JSON文件是否存在
    if not json_file.exists():
        print(f"❌ JSON文件不存在: {json_file}")
        return False

    # 加载JSON数据
    print(f"📖 正在读取: {json_file}")
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 创建Excel writer
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # 创建概览页
        overview_data = {
            '项目': ['版本', '最后更新', '说明'],
            '值': [
                data.get('version', ''),
                data.get('last_updated', ''),
                data.get('description', '')
            ]
        }
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='概览', index=False)

        # 为每个测试套件创建一个sheet
        for suite_name, suite_data in data['test_suites'].items():
            # 转换为DataFrame
            df = pd.DataFrame(suite_data['test_cases'])

            # 写入Excel
            sheet_name = suite_name[:31]  # Excel sheet名称限制31字符
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"  ✓ 已导出测试套件: {suite_name} ({len(df)} 个用例)")

        # 导出示例数据
        if 'sample_data' in data:
            sample_items = []
            for key, value in data['sample_data']['data'].items():
                sample_items.append({'字段': key, '值': value})

            sample_df = pd.DataFrame(sample_items)
            sample_df.to_excel(writer, sheet_name='示例数据', index=False)
            print(f"  ✓ 已导出示例数据")

    # 美化Excel
    print("🎨 正在美化Excel...")
    beautify_excel(excel_file)

    print(f"✅ 导出成功: {excel_file}")
    print(f"📊 文件大小: {excel_file.stat().st_size / 1024:.1f} KB")
    return True


def beautify_excel(excel_file):
    """美化Excel格式"""
    wb = load_workbook(excel_file)

    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 单元格样式
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置标题行样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置数据单元格样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_alignment

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # 设置列宽（限制最大宽度）
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 设置行高
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 30

    wb.save(excel_file)


if __name__ == "__main__":
    print("=" * 60)
    print("📤 JSON测试用例导出工具")
    print("=" * 60)
    print()

    success = export_to_excel()

    print()
    print("=" * 60)
    if success:
        print("✅ 导出完成！")
        print()
        print("📝 使用提示:")
        print("  1. 用Excel打开文件查看测试用例")
        print("  2. 可以在Excel中编辑测试用例")
        print("  3. 编辑完成后使用 import_from_excel.py 导入回JSON")
    else:
        print("❌ 导出失败！")
        sys.exit(1)
    print("=" * 60)
