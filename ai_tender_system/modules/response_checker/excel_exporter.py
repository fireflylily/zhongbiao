#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
应答文件自检Excel导出器

导出清单式检查报告
"""

import io
from typing import Dict, Any, List
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


class ResponseCheckExcelExporter:
    """
    应答文件自检Excel导出器

    导出结构：
    - Sheet1: 检查汇总
    - Sheet2: 检查明细
    - Sheet3: 不符合项汇总
    """

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if HAS_OPENPYXL else None
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if HAS_OPENPYXL else None
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if HAS_OPENPYXL else None
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if HAS_OPENPYXL else None

    def __init__(self):
        if not HAS_OPENPYXL:
            logger.warning("openpyxl未安装，Excel导出功能不可用")

    def export(self, result: Dict[str, Any]) -> io.BytesIO:
        """
        导出检查结果为Excel

        Args:
            result: 检查结果字典

        Returns:
            Excel文件的BytesIO对象
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl未安装，无法导出Excel")

        wb = Workbook()

        # Sheet1: 检查汇总
        self._create_summary_sheet(wb, result)

        # Sheet2: 检查明细
        self._create_detail_sheet(wb, result)

        # Sheet3: 不符合项汇总
        self._create_failed_sheet(wb, result)

        # 删除默认的空sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _create_summary_sheet(self, wb: Workbook, result: Dict[str, Any]):
        """创建检查汇总Sheet"""
        ws = wb.create_sheet("检查汇总", 0)

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50

        # 标题
        ws.merge_cells('A1:B1')
        ws['A1'] = "应答文件自检报告"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # 基本信息
        row = 3
        info_items = [
            ("文件名", result.get('file_name', '')),
            ("检查时间", result.get('check_time', '')),
            ("总检查项", str(result.get('statistics', {}).get('total_items', 0))),
            ("符合项", str(result.get('statistics', {}).get('pass_count', 0))),
            ("不符合项", str(result.get('statistics', {}).get('fail_count', 0))),
            ("无法判断项", str(result.get('statistics', {}).get('unknown_count', 0))),
            ("分析耗时", f"{result.get('analysis_time', 0):.2f}秒"),
            ("文档页数", str(result.get('total_pages', 0))),
        ]

        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # 各类别汇总
        row += 1
        ws[f'A{row}'] = "检查类别汇总"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        # 表头
        headers = ['类别名称', '符合', '不符合', '无法判断', '状态']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # 类别数据
        categories = result.get('categories', [])
        for cat in categories:
            ws.cell(row=row, column=1, value=cat.get('category_name', ''))
            ws.cell(row=row, column=2, value=cat.get('pass_count', 0))
            ws.cell(row=row, column=3, value=cat.get('fail_count', 0))
            ws.cell(row=row, column=4, value=cat.get('unknown_count', 0))

            # 状态
            status_icon = cat.get('status_icon', '')
            ws.cell(row=row, column=5, value=status_icon)
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')

            # 根据状态设置背景色
            if cat.get('fail_count', 0) > 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.FAIL_FILL
            elif cat.get('unknown_count', 0) > 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.WARNING_FILL
            else:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.PASS_FILL

            row += 1

    def _create_detail_sheet(self, wb: Workbook, result: Dict[str, Any]):
        """创建检查明细Sheet"""
        ws = wb.create_sheet("检查明细")

        # 设置列宽
        col_widths = [6, 15, 30, 10, 30, 10, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 表头
        headers = ['序号', '类别', '检查项', '状态', '说明', '位置', '建议']
        row = 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据行
        row = 2
        seq = 1
        categories = result.get('categories', [])

        for cat in categories:
            items = cat.get('items', [])
            for item in items:
                ws.cell(row=row, column=1, value=seq)
                ws.cell(row=row, column=2, value=cat.get('category_name', ''))
                ws.cell(row=row, column=3, value=item.get('name', ''))

                status = item.get('status', '')
                ws.cell(row=row, column=4, value=status)

                ws.cell(row=row, column=5, value=item.get('detail', ''))
                ws.cell(row=row, column=6, value=item.get('location', ''))
                ws.cell(row=row, column=7, value=item.get('suggestion', ''))

                # 根据状态设置行背景色
                if status == '不符合':
                    fill = self.FAIL_FILL
                elif status == '无法判断':
                    fill = self.WARNING_FILL
                else:
                    fill = self.PASS_FILL

                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)

                row += 1
                seq += 1

    def _create_failed_sheet(self, wb: Workbook, result: Dict[str, Any]):
        """创建不符合项汇总Sheet"""
        ws = wb.create_sheet("不符合项汇总")

        # 设置列宽
        col_widths = [6, 15, 30, 30, 10, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 表头
        headers = ['序号', '类别', '检查项', '问题说明', '位置', '修改建议']
        row = 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 收集不符合项
        row = 2
        seq = 1
        categories = result.get('categories', [])

        for cat in categories:
            items = cat.get('items', [])
            for item in items:
                if item.get('status') == '不符合':
                    ws.cell(row=row, column=1, value=seq)
                    ws.cell(row=row, column=2, value=cat.get('category_name', ''))
                    ws.cell(row=row, column=3, value=item.get('name', ''))
                    ws.cell(row=row, column=4, value=item.get('detail', ''))
                    ws.cell(row=row, column=5, value=item.get('location', ''))
                    ws.cell(row=row, column=6, value=item.get('suggestion', ''))

                    # 设置行样式
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = self.FAIL_FILL
                        ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)

                    row += 1
                    seq += 1

        # 如果没有不符合项，显示提示
        if seq == 1:
            ws.merge_cells('A2:F2')
            ws['A2'] = "恭喜！未发现不符合项"
            ws['A2'].font = Font(size=12, color="008000")
            ws['A2'].alignment = Alignment(horizontal='center')
            ws['A2'].fill = self.PASS_FILL
