#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 导出器 5.0 版本 - 增强版风险分析报告导出

新增字段：
- 风险等级（带颜色）
- 招标文件原文
- 建议避坑动作(Todo)
- 应答自检结论
- 问题深度解析
- 位置索引

新增 Sheet：
- 汇总页
- 风险明细页
- 双向对账页（如有）
"""

import io
from typing import List, Dict, Optional
from datetime import datetime
from pathlib import Path

import sys
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from common.logger import get_module_logger

# 尝试导入 xlsxwriter，如果没有则使用 openpyxl
try:
    import xlsxwriter
    USE_XLSXWRITER = True
except ImportError:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    USE_XLSXWRITER = False

from .schemas import RiskItem, RiskAnalysisResult, ReconcileResult

logger = get_module_logger("risk_analyzer.excel_exporter")


class ExcelExporterV5:
    """
    增强版 Excel 导出器

    特性：
    - 多 Sheet 结构（汇总、明细、对账）
    - 风险等级颜色编码
    - 合规状态颜色编码
    - 第一行汇总摘要
    """

    # 风险等级颜色
    RISK_COLORS = {
        'high': '#FFCDD2',      # 红色
        'medium': '#FFE0B2',    # 橙色
        'low': '#E3F2FD'        # 蓝色
    }

    # 合规状态颜色
    COMPLIANCE_COLORS = {
        'compliant': '#C8E6C9',      # 绿色
        'non_compliant': '#FFCDD2',  # 红色
        'partial': '#FFF9C4',        # 黄色
        'unknown': '#EEEEEE'         # 灰色
    }

    # 风险等级文本
    RISK_LEVEL_TEXT = {
        'high': '高风险',
        'medium': '中风险',
        'low': '低风险'
    }

    # 合规状态文本
    COMPLIANCE_STATUS_TEXT = {
        'compliant': '已符合',
        'non_compliant': '不符合',
        'partial': '部分符合',
        'unknown': '待检查'
    }

    def __init__(self):
        """初始化导出器"""
        logger.info(f"Excel 导出器初始化, 使用 {'xlsxwriter' if USE_XLSXWRITER else 'openpyxl'}")

    def export(self,
               result: RiskAnalysisResult,
               file_name: str = '',
               include_reconcile: bool = True) -> io.BytesIO:
        """
        导出增强版 Excel

        Args:
            result: 分析结果
            file_name: 文件名（用于显示）
            include_reconcile: 是否包含对账页

        Returns:
            BytesIO: Excel 文件内容
        """
        output = io.BytesIO()

        if USE_XLSXWRITER:
            return self._export_with_xlsxwriter(result, file_name, include_reconcile, output)
        else:
            return self._export_with_openpyxl(result, file_name, include_reconcile, output)

    def _export_with_xlsxwriter(self,
                                 result: RiskAnalysisResult,
                                 file_name: str,
                                 include_reconcile: bool,
                                 output: io.BytesIO) -> io.BytesIO:
        """使用 xlsxwriter 导出"""
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        # 创建格式
        formats = self._create_formats_xlsxwriter(workbook)

        # Sheet 1: 汇总
        self._build_summary_sheet_xlsxwriter(workbook, formats, result, file_name)

        # Sheet 2: 风险明细
        self._build_detail_sheet_xlsxwriter(workbook, formats, result.risk_items)

        # Sheet 3: 双向对账（如有）
        if include_reconcile and result.reconcile_results:
            self._build_reconcile_sheet_xlsxwriter(workbook, formats, result)

        workbook.close()
        output.seek(0)
        return output

    def _create_formats_xlsxwriter(self, workbook):
        """创建 xlsxwriter 格式"""
        return {
            'header': workbook.add_format({
                'bold': True,
                'bg_color': '#1976D2',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True
            }),
            'cell': workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'top',
                'text_wrap': True
            }),
            'cell_center': workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            }),
            'title': workbook.add_format({
                'bold': True,
                'font_size': 14,
                'align': 'left'
            }),
            'summary_text': workbook.add_format({
                'bold': True,
                'font_size': 11,
                'text_wrap': True
            }),
            'risk_high': workbook.add_format({
                'border': 1,
                'bg_color': '#FFCDD2',
                'align': 'center',
                'valign': 'vcenter'
            }),
            'risk_medium': workbook.add_format({
                'border': 1,
                'bg_color': '#FFE0B2',
                'align': 'center',
                'valign': 'vcenter'
            }),
            'risk_low': workbook.add_format({
                'border': 1,
                'bg_color': '#E3F2FD',
                'align': 'center',
                'valign': 'vcenter'
            }),
            'compliant': workbook.add_format({
                'border': 1,
                'bg_color': '#C8E6C9',
                'align': 'center',
                'valign': 'vcenter'
            }),
            'non_compliant': workbook.add_format({
                'border': 1,
                'bg_color': '#FFCDD2',
                'align': 'center',
                'valign': 'vcenter'
            }),
            'partial': workbook.add_format({
                'border': 1,
                'bg_color': '#FFF9C4',
                'align': 'center',
                'valign': 'vcenter'
            }),
            'unknown': workbook.add_format({
                'border': 1,
                'bg_color': '#EEEEEE',
                'align': 'center',
                'valign': 'vcenter'
            })
        }

    def _build_summary_sheet_xlsxwriter(self, workbook, formats, result, file_name):
        """构建汇总页 (xlsxwriter)"""
        sheet = workbook.add_worksheet('汇总')

        # 统计数据
        stats = result.get_statistics()
        reconcile_summary = result.get_reconcile_summary()

        # 标题
        sheet.write('A1', '标书宝 5.0 风险分析报告', formats['title'])
        sheet.write('A2', f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}')

        # 基本信息
        sheet.write('A4', '招标文件', formats['header'])
        sheet.write('B4', file_name or '未知', formats['cell'])

        sheet.write('A5', '分析耗时', formats['header'])
        sheet.write('B5', f'{result.analysis_time:.1f} 秒', formats['cell'])

        sheet.write('A6', '目录识别', formats['header'])
        sheet.write('B6', '是' if result.has_toc else '否（智能降级）', formats['cell'])

        # 风险统计
        sheet.write('A8', '风险统计', formats['header'])
        sheet.merge_range('B8:C8', '', formats['header'])

        sheet.write('A9', '总计', formats['cell_center'])
        sheet.write('B9', stats['total_items'], formats['cell_center'])

        sheet.write('A10', '🔴 高风险', formats['risk_high'])
        sheet.write('B10', stats['high_risk_count'], formats['risk_high'])

        sheet.write('A11', '🟡 中风险', formats['risk_medium'])
        sheet.write('B11', stats['medium_risk_count'], formats['risk_medium'])

        sheet.write('A12', '🔵 低风险', formats['risk_low'])
        sheet.write('B12', stats['low_risk_count'], formats['risk_low'])

        # 汇总摘要
        sheet.write('A14', '分析摘要', formats['header'])
        sheet.merge_range('A15:D16', result.summary, formats['summary_text'])

        # 对账汇总（如有）
        if reconcile_summary:
            sheet.write('A18', '对账统计', formats['header'])
            sheet.merge_range('B18:C18', '', formats['header'])

            sheet.write('A19', '已检查', formats['cell_center'])
            sheet.write('B19', reconcile_summary.get('total_checked', 0), formats['cell_center'])

            sheet.write('A20', '🟢 已符合', formats['compliant'])
            sheet.write('B20', reconcile_summary.get('compliant', 0), formats['compliant'])

            sheet.write('A21', '🔴 不符合', formats['non_compliant'])
            sheet.write('B21', reconcile_summary.get('non_compliant', 0), formats['non_compliant'])

            sheet.write('A22', '🟡 部分符合', formats['partial'])
            sheet.write('B22', reconcile_summary.get('partial', 0), formats['partial'])

        # 设置列宽
        sheet.set_column('A:A', 15)
        sheet.set_column('B:B', 50)
        sheet.set_column('C:D', 20)

    def _build_detail_sheet_xlsxwriter(self, workbook, formats, risk_items):
        """构建风险明细页 (xlsxwriter)"""
        sheet = workbook.add_worksheet('风险明细')

        # 表头 - 5.0 新增字段
        headers = [
            '序号', '风险等级', '风险类型', '章节来源', '位置索引',
            '条款位置', '招标文件原文', '核心要求', '问题深度解析',
            '建议避坑动作(Todo)', '应答自检结论'
        ]

        for col, header in enumerate(headers):
            sheet.write(0, col, header, formats['header'])

        # 写入数据
        for row, item in enumerate(risk_items, start=1):
            sheet.write(row, 0, row, formats['cell_center'])

            # 风险等级（带颜色）
            level = item.risk_level or 'medium'
            level_format = formats.get(f'risk_{level}', formats['cell_center'])
            level_text = self.RISK_LEVEL_TEXT.get(level, '中风险')
            sheet.write(row, 1, level_text, level_format)

            sheet.write(row, 2, item.risk_type or '-', formats['cell'])
            sheet.write(row, 3, f"第{item.source_chunk + 1}部分", formats['cell'])
            sheet.write(row, 4, item.position_index or '-', formats['cell'])
            sheet.write(row, 5, item.location or '-', formats['cell'])
            sheet.write(row, 6, item.original_text or '-', formats['cell'])
            sheet.write(row, 7, item.requirement or '-', formats['cell'])
            sheet.write(row, 8, item.deep_analysis or '-', formats['cell'])
            sheet.write(row, 9, item.todo_action or item.suggestion or '-', formats['cell'])

            # 应答自检结论（带颜色）
            status = item.compliance_status or 'unknown'
            status_format = formats.get(status, formats['cell_center'])
            status_text = self.COMPLIANCE_STATUS_TEXT.get(status, '待检查')
            sheet.write(row, 10, status_text, status_format)

        # 设置列宽
        col_widths = [6, 10, 12, 12, 12, 20, 40, 40, 40, 40, 12]
        for col, width in enumerate(col_widths):
            sheet.set_column(col, col, width)

    def _build_reconcile_sheet_xlsxwriter(self, workbook, formats, result):
        """构建双向对账页 (xlsxwriter)"""
        sheet = workbook.add_worksheet('双向对账')

        headers = [
            '序号', '招标要求', '应答内容', '匹配度',
            '合规状态', '问题描述', '修复建议'
        ]

        for col, header in enumerate(headers):
            sheet.write(0, col, header, formats['header'])

        row = 1
        for i, rec in enumerate(result.reconcile_results):
            sheet.write(row, 0, i + 1, formats['cell_center'])
            sheet.write(row, 1, rec.bid_requirement or '-', formats['cell'])
            sheet.write(row, 2, rec.response_content or '-', formats['cell'])
            sheet.write(row, 3, f"{rec.match_score * 100:.0f}%", formats['cell_center'])

            status = rec.compliance_status or 'unknown'
            status_format = formats.get(status, formats['cell_center'])
            status_text = self.COMPLIANCE_STATUS_TEXT.get(status, '未知')
            sheet.write(row, 4, status_text, status_format)

            sheet.write(row, 5, rec.overall_assessment or '-', formats['cell'])
            sheet.write(row, 6, rec.fix_suggestion or '-', formats['cell'])
            row += 1

        # 设置列宽
        col_widths = [6, 40, 40, 10, 12, 40, 40]
        for col, width in enumerate(col_widths):
            sheet.set_column(col, col, width)

    def _export_with_openpyxl(self,
                               result: RiskAnalysisResult,
                               file_name: str,
                               include_reconcile: bool,
                               output: io.BytesIO) -> io.BytesIO:
        """使用 openpyxl 导出（备用方案）"""
        workbook = openpyxl.Workbook()

        # 移除默认 sheet
        default_sheet = workbook.active
        default_sheet.title = '汇总'

        # Sheet 1: 汇总
        self._build_summary_sheet_openpyxl(default_sheet, result, file_name)

        # Sheet 2: 风险明细
        detail_sheet = workbook.create_sheet('风险明细')
        self._build_detail_sheet_openpyxl(detail_sheet, result.risk_items)

        # Sheet 3: 双向对账（如有）
        if include_reconcile and result.reconcile_results:
            reconcile_sheet = workbook.create_sheet('双向对账')
            self._build_reconcile_sheet_openpyxl(reconcile_sheet, result)

        workbook.save(output)
        output.seek(0)
        return output

    def _build_summary_sheet_openpyxl(self, sheet, result, file_name):
        """构建汇总页 (openpyxl)"""
        stats = result.get_statistics()

        sheet['A1'] = '标书宝 5.0 风险分析报告'
        sheet['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        sheet['A4'] = '招标文件'
        sheet['B4'] = file_name or '未知'

        sheet['A5'] = '分析耗时'
        sheet['B5'] = f'{result.analysis_time:.1f} 秒'

        sheet['A7'] = '风险统计'
        sheet['A8'] = '总计'
        sheet['B8'] = stats['total_items']

        sheet['A9'] = '高风险'
        sheet['B9'] = stats['high_risk_count']

        sheet['A10'] = '中风险'
        sheet['B10'] = stats['medium_risk_count']

        sheet['A11'] = '低风险'
        sheet['B11'] = stats['low_risk_count']

        sheet['A13'] = '分析摘要'
        sheet['A14'] = result.summary

        # 设置列宽
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 50

    def _build_detail_sheet_openpyxl(self, sheet, risk_items):
        """构建风险明细页 (openpyxl)"""
        headers = [
            '序号', '风险等级', '风险类型', '章节来源', '位置索引',
            '条款位置', '招标文件原文', '核心要求', '问题深度解析',
            '建议避坑动作(Todo)', '应答自检结论'
        ]

        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row, item in enumerate(risk_items, start=2):
            sheet.cell(row=row, column=1, value=row - 1)
            sheet.cell(row=row, column=2, value=self.RISK_LEVEL_TEXT.get(item.risk_level, '中风险'))
            sheet.cell(row=row, column=3, value=item.risk_type or '-')
            sheet.cell(row=row, column=4, value=f"第{item.source_chunk + 1}部分")
            sheet.cell(row=row, column=5, value=item.position_index or '-')
            sheet.cell(row=row, column=6, value=item.location or '-')
            sheet.cell(row=row, column=7, value=item.original_text or '-')
            sheet.cell(row=row, column=8, value=item.requirement or '-')
            sheet.cell(row=row, column=9, value=item.deep_analysis or '-')
            sheet.cell(row=row, column=10, value=item.todo_action or item.suggestion or '-')
            sheet.cell(row=row, column=11, value=self.COMPLIANCE_STATUS_TEXT.get(item.compliance_status, '待检查'))

    def _build_reconcile_sheet_openpyxl(self, sheet, result):
        """构建双向对账页 (openpyxl)"""
        headers = ['序号', '招标要求', '应答内容', '匹配度', '合规状态', '问题描述', '修复建议']

        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row, rec in enumerate(result.reconcile_results, start=2):
            sheet.cell(row=row, column=1, value=row - 1)
            sheet.cell(row=row, column=2, value=rec.bid_requirement or '-')
            sheet.cell(row=row, column=3, value=rec.response_content or '-')
            sheet.cell(row=row, column=4, value=f"{rec.match_score * 100:.0f}%")
            sheet.cell(row=row, column=5, value=self.COMPLIANCE_STATUS_TEXT.get(rec.compliance_status, '未知'))
            sheet.cell(row=row, column=6, value=rec.overall_assessment or '-')
            sheet.cell(row=row, column=7, value=rec.fix_suggestion or '-')


# 便捷函数
def export_risk_report(result: RiskAnalysisResult,
                       file_name: str = '',
                       include_reconcile: bool = True) -> io.BytesIO:
    """
    便捷函数：导出风险分析报告

    Args:
        result: 分析结果
        file_name: 文件名
        include_reconcile: 是否包含对账

    Returns:
        BytesIO: Excel 文件内容
    """
    exporter = ExcelExporterV5()
    return exporter.export(result, file_name, include_reconcile)
